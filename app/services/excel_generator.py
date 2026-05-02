import re
import pandas as pd
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


COR_ROXO = "5B2C83"
COR_AMARELO = "FFD966"
COR_VERDE = "D9EAD3"
COR_CINZA = "F2F2F2"
COR_BRANCO = "FFFFFF"


def moeda(valor):
    try:
        return float(valor)
    except Exception:
        return 0.0


def limpar_nome_cliente(detalhes):
    if not detalhes:
        return ""

    partes = [p.strip() for p in str(detalhes).split("|") if p.strip()]

    remover = [
        "Recebimento Pix",
        "Pagamento Pix",
        "Transferência Pix",
        "MHM CAIXAS COMERCIO DE EMBALAGENS LTDA",
    ]

    candidatos = []

    for p in partes:
        pu = p.upper()

        if any(r.upper() in pu for r in remover):
            continue

        if re.search(r"\d{2}\.\d{3}\.\d{3}", p):
            continue

        if "***" in p:
            continue

        candidatos.append(p)

    return candidatos[0] if candidatos else ""


def limpar_nome_fornecedor(texto):
    if not texto:
        return "Sem fornecedor"

    t = str(texto).upper()

    t = t.replace("FORNECEDOR", "")
    t = t.replace("FAV.:", "")
    t = t.replace("FAV:", "")

    t = re.sub(r"\bNF\b.*", "", t)
    t = re.sub(r"\bRM\b.*", "", t)
    t = re.sub(r"\bRMS\b.*", "", t)
    t = re.sub(r"\bNOTA\b.*", "", t)
    t = re.sub(r"\bNFE\b.*", "", t)

    t = re.sub(r"\d+", "", t)
    t = re.sub(r"[-–—_/.,]", " ", t)
    t = re.sub(r"\s+", " ", t)

    t = t.strip()

    correcoes = {
        "CYCLOPACK": "Cyclopack",
        "TECNIPACK": "Tecnipack",
        "TECHNIPACK": "Tecnipack",
        "NOVAPLASTIC": "Novaplastic",
        "F WEBPACK": "F Webpack",
        "WEBPACK": "Webpack",
        "UNIPEL": "Unipel",
        "ALLTAPE": "Alltape",
        "NOVABOLHA": "Novabolha",
        "PLASTBOLHA": "Plastbolha",
        "IDEAL": "Ideal",
        "PREMISSE": "Premisse",
        "SIMPRA": "Simpra",
        "TECNICAIXA": "Tecnicaixa",
        "PSR": "Psr",
        "ISOEP": "Isoep",
        "MARRECO": "Marreco",
        "ECONERG": "Econerg",
        "AUXILIADORA": "Auxiliadora",
        "INDUSPAPER": "Induspaper",
        "FITACHINESA": "Fita Chinesa",
        "HF AMBIENTAL": "Hf Ambiental",
        "SHEERPACK": "Sheerpack",
        "PROPAPACK": "Propapack",
        "TALGE": "Talge",
        "SUL MANTIQUEIRA": "Sul Mantiqueira",
        "LUPMA": "Lupma",
    }

    for chave, nome in correcoes.items():
        if chave in t:
            return nome

    if not t:
        return "Sem fornecedor"

    return t.title()


def tipo_movimento(descricao, detalhes):
    texto = f"{descricao} {detalhes}".upper()

    if "ANTECIPA" in texto:
        return "ANTECIPAÇÃO"

    if "PIX" in texto:
        return "PIX"

    if "MAESTRO" in texto or "VISA ELECTRON" in texto or "ELO DÉBITO" in texto or "DEB._" in texto:
        return "DÉBITO"

    if "MASTERCARD" in texto or "VISA" in texto or "CRED" in texto or "CRÉD" in texto:
        return "CRÉDITO"

    if "BOLETO" in texto or "COBRANÇA" in texto or "COBRANCA" in texto:
        return "BOLETO"

    if "DINHEIRO" in texto:
        return "DINHEIRO"

    return descricao


def ordem_movimento(descricao, detalhes):
    texto = f"{descricao} {detalhes}".upper()

    if "PIX" in texto:
        return 1

    if "MAESTRO" in texto or "VISA ELECTRON" in texto or "ELO DÉBITO" in texto or "DEB._" in texto:
        return 2

    if "CRED" in texto or "CRÉD" in texto or "MASTERCARD" in texto or "VISA" in texto:
        return 3

    if "BOLETO" in texto or "COBRANÇA" in texto or "COBRANCA" in texto:
        return 4

    if "ANTECIPA" in texto:
        return 5

    return 6


def preparar_transacoes(transacoes):
    df = pd.DataFrame(transacoes)

    if df.empty:
        return df

    df["valor"] = pd.to_numeric(df["valor"], errors="coerce").fillna(0)

    df["Data"] = pd.to_datetime(
        df["data"],
        dayfirst=True,
        errors="coerce"
    ).dt.strftime("%d/%m/%Y")

    df["Descrição"] = df.apply(
        lambda r: tipo_movimento(r.get("descricao", ""), r.get("detalhes", "")),
        axis=1
    )

    df["Detalhes"] = df["detalhes"].apply(limpar_nome_cliente)
    df["Valor"] = df["valor"]
    df["TipoInterno"] = df["tipo"]
    df["CategoriaInterna"] = df["categoria"]

    df["_ordem"] = df.apply(
        lambda r: ordem_movimento(r.get("descricao", ""), r.get("detalhes", "")),
        axis=1
    )

    df = df.sort_values(by=["_ordem", "Data"], ascending=[True, True])

    return df[
        [
            "Data",
            "Descrição",
            "Detalhes",
            "Valor",
            "TipoInterno",
            "CategoriaInterna",
        ]
    ]


def aplicar_estilo_tabela(ws):
    header_fill = PatternFill("solid", fgColor=COR_ROXO)
    header_font = Font(color=COR_BRANCO, bold=True)

    border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            value = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(value))

        ws.column_dimensions[col_letter].width = min(max_len + 4, 55)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws["D"][1:]:
        cell.number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'


def escrever_secao(ws, col_nome, col_valor, linha, titulo, dados, cor=COR_CINZA):
    ws.cell(row=linha, column=col_nome).value = titulo
    ws.cell(row=linha, column=col_nome).font = Font(bold=True)
    ws.cell(row=linha, column=col_nome).fill = PatternFill("solid", fgColor=cor)

    ws.cell(row=linha, column=col_valor).value = "VALOR"
    ws.cell(row=linha, column=col_valor).font = Font(bold=True)
    ws.cell(row=linha, column=col_valor).fill = PatternFill("solid", fgColor=cor)

    linha += 1
    total = 0

    for nome, valor in dados:
        ws.cell(row=linha, column=col_nome).value = nome or "SEM DETALHE"
        ws.cell(row=linha, column=col_valor).value = moeda(valor)
        ws.cell(row=linha, column=col_valor).number_format = 'R$ #,##0.00'
        total += moeda(valor)
        linha += 1

    linha += 1

    ws.cell(row=linha, column=col_nome).value = "TOTAL"
    ws.cell(row=linha, column=col_nome).font = Font(bold=True)

    ws.cell(row=linha, column=col_valor).value = total
    ws.cell(row=linha, column=col_valor).font = Font(bold=True)
    ws.cell(row=linha, column=col_valor).number_format = 'R$ #,##0.00'

    return linha + 2


def resumo_por_texto(df, categorias=None, entrada=None, normalizar_fornecedor=False):
    base = df.copy()

    if categorias:
        cats = [c.upper() for c in categorias]
        base = base[base["CategoriaInterna"].str.upper().isin(cats)]

    if entrada is True:
        base = base[base["Valor"] > 0]

    if entrada is False:
        base = base[base["Valor"] < 0]

    if base.empty:
        return []

    if normalizar_fornecedor:
        base["NomeResumo"] = base["Detalhes"].apply(limpar_nome_fornecedor)
    else:
        base["NomeResumo"] = base["Detalhes"]

    resumo = base.groupby("NomeResumo")["Valor"].sum().reset_index()
    resumo["ValorAbs"] = resumo["Valor"].abs()
    resumo = resumo.sort_values("ValorAbs", ascending=False)

    return [
        (r["NomeResumo"] or "SEM DETALHE", r["ValorAbs"])
        for _, r in resumo.iterrows()
    ]


def criar_graficos(ws, metricas):
    ws.title = "Métricas"

    ws["A1"] = "MÉTRICAS PARA GRÁFICOS"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].fill = PatternFill("solid", fgColor=COR_AMARELO)

    ws["A3"] = "Métrica"
    ws["B3"] = "Valor"

    for cell in ws[3]:
        cell.font = Font(bold=True, color=COR_BRANCO)
        cell.fill = PatternFill("solid", fgColor=COR_ROXO)

    linha = 4
    for nome, valor in metricas:
        ws.cell(row=linha, column=1).value = nome
        ws.cell(row=linha, column=2).value = moeda(valor)
        ws.cell(row=linha, column=2).number_format = 'R$ #,##0.00'
        linha += 1

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18

    pie = PieChart()
    pie.title = "Distribuição Financeira"
    data = Reference(ws, min_col=2, min_row=4, max_row=linha - 1)
    labels = Reference(ws, min_col=1, min_row=4, max_row=linha - 1)
    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)
    pie.height = 12
    pie.width = 16
    ws.add_chart(pie, "D3")

    bar = BarChart()
    bar.title = "Resumo por Métrica"
    bar.y_axis.title = "Valor"
    bar.x_axis.title = "Métrica"
    bar.add_data(data, titles_from_data=False)
    bar.set_categories(labels)
    bar.height = 10
    bar.width = 18
    ws.add_chart(bar, "D25")


def criar_resumo(ws, df):
    ws.title = "Sicoob Resumo"

    ws.merge_cells("A1:H1")
    ws["A1"] = "DRE SICOOB"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].fill = PatternFill("solid", fgColor=COR_AMARELO)

    entradas_vendas = [
        ("BOLETOS", df[(df["Valor"] > 0) & (df["Descrição"].str.upper().str.contains("BOLETO|COBRANÇA|COBRANCA", na=False))]["Valor"].sum()),
        ("CARTÕES", df[(df["Valor"] > 0) & (df["Descrição"].str.upper().isin(["DÉBITO", "CRÉDITO"]))]["Valor"].sum()),
        ("PIX RECEBIDO", df[(df["Valor"] > 0) & (df["Descrição"].str.upper() == "PIX")]["Valor"].sum()),
    ]

    cartoes = [
        ("CARTÃO DÉBITO", df[(df["Valor"] > 0) & (df["Descrição"].str.upper() == "DÉBITO")]["Valor"].sum()),
        ("CARTÃO CRÉDITO", df[(df["Valor"] > 0) & (df["Descrição"].str.upper() == "CRÉDITO")]["Valor"].sum()),
    ]

    fornecedores = resumo_por_texto(df, categorias=["Fornecedores"], entrada=False, normalizar_fornecedor=True)
    impostos = resumo_por_texto(df, categorias=["Impostos"], entrada=False)
    tarifas = resumo_por_texto(df, categorias=["Taxas Bancárias"], entrada=False)
    lucros = resumo_por_texto(df, categorias=["Distribuição de Lucros"], entrada=False)
    folha = resumo_por_texto(df, categorias=["Salários"], entrada=False)

    despesas_fixas = resumo_por_texto(df, categorias=["Despesas"], entrada=False)[:35]
    despesas_variaveis = resumo_por_texto(df, categorias=["Despesas"], entrada=False)[35:]

    adiantamentos = [
        (
            "Salários Agosto",
            abs(
                df[
                    (df["Valor"] < 0)
                    & (
                        df["Detalhes"].str.upper().str.contains("ADIANTAMENTO|ADIANT|SALARIO|SALÁRIO|FERIAS|FÉRIAS", na=False)
                        | df["CategoriaInterna"].str.upper().str.contains("SALÁRIO|SALARIOS|SALÁRIOS", na=False)
                    )
                ]["Valor"].sum()
            )
        )
    ]

    linha_a = 3
    linha_d = 3
    linha_g = 3

    linha_a = escrever_secao(ws, 1, 2, linha_a, "ENTRADAS (VENDAS)", entradas_vendas, COR_VERDE)
    linha_a = escrever_secao(ws, 1, 2, linha_a, "CARTÕES DETALHADOS", cartoes, COR_CINZA)

    linha_d = escrever_secao(ws, 4, 5, linha_d, "DESPESAS FIXAS", despesas_fixas, COR_CINZA)
    linha_d = escrever_secao(ws, 4, 5, linha_d, "DESPESAS VARIÁVEIS", despesas_variaveis, COR_CINZA)
    linha_d = escrever_secao(ws, 4, 5, linha_d, "FOLHA DE PAGAMENTO", folha, COR_CINZA)
    linha_d = escrever_secao(ws, 4, 5, linha_d, "ADIANTAMENTOS", adiantamentos, COR_CINZA)

    linha_g = escrever_secao(ws, 7, 8, linha_g, "FORNECEDORES", fornecedores, COR_CINZA)
    linha_g = escrever_secao(ws, 7, 8, linha_g, "IMPOSTOS", impostos, COR_CINZA)
    linha_g = escrever_secao(ws, 7, 8, linha_g, "TARIFAS BANCÁRIAS", tarifas, COR_CINZA)
    linha_g = escrever_secao(ws, 7, 8, linha_g, "DISTRIBUIÇÃO DE LUCROS", lucros, COR_CINZA)

    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 26

    border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center")
            cell.border = border

    total_fornecedores = sum(v for _, v in fornecedores)
    total_despesas_fixas = sum(v for _, v in despesas_fixas)
    total_despesas_variaveis = sum(v for _, v in despesas_variaveis)
    total_impostos = sum(v for _, v in impostos)
    total_tarifas = sum(v for _, v in tarifas)
    total_folha = sum(v for _, v in folha)
    total_lucros = sum(v for _, v in lucros)
    total_debito = cartoes[0][1]
    total_credito = cartoes[1][1]

    metricas = [
        ("Fornecedores", total_fornecedores),
        ("Despesas Fixas", total_despesas_fixas),
        ("Despesas Variáveis", total_despesas_variaveis),
        ("Impostos", total_impostos),
        ("Tarifas Bancárias", total_tarifas),
        ("Folha de Pagamento", total_folha),
        ("Distribuição de Lucros", total_lucros),
        ("Cartão Débito", total_debito),
        ("Cartão Crédito", total_credito),
    ]

    return metricas


def gerar_excel(transacoes, file_path_saida):
    df_interno = preparar_transacoes(transacoes)

    if df_interno.empty:
        return None

    df_sicoob = df_interno[
        [
            "Data",
            "Descrição",
            "Detalhes",
            "Valor",
        ]
    ]

    with pd.ExcelWriter(file_path_saida, engine="openpyxl") as writer:
        df_sicoob.to_excel(writer, sheet_name="Sicoob", index=False)

        workbook = writer.book

        ws_sicoob = workbook["Sicoob"]
        aplicar_estilo_tabela(ws_sicoob)

        ws_resumo = workbook.create_sheet("Sicoob Resumo")
        metricas = criar_resumo(ws_resumo, df_interno)

        ws_metricas = workbook.create_sheet("Métricas")
        criar_graficos(ws_metricas, metricas)

    return file_path_saida
